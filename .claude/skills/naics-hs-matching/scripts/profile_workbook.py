from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def compact(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value[:250]
    return value


def value_type(value: Any) -> str:
    if value is None:
        return "blank"
    return type(value).__name__


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: profile_workbook.py <input.xlsx> <output.json>")

    input_path = Path(sys.argv[1]).expanduser()
    output_path = Path(sys.argv[2])
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    profile: dict[str, Any] = {
        "input_path": str(input_path),
        "sheets": [],
    }

    for ws in wb.worksheets:
        first_rows = []
        type_counts_by_col: dict[int, Counter[str]] = defaultdict(Counter)
        nonblank_by_col: Counter[int] = Counter()
        sample_limit = min(ws.max_row or 0, 5000)

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx <= 10:
                first_rows.append([compact(cell) for cell in row])
            if row_idx <= sample_limit:
                for col_idx, cell in enumerate(row, start=1):
                    type_counts_by_col[col_idx][value_type(cell)] += 1
                    if cell not in (None, ""):
                        nonblank_by_col[col_idx] += 1
            if row_idx >= max(10, sample_limit):
                break

        header = first_rows[0] if first_rows else []
        normalized_headers = [
            str(cell).strip() if cell is not None else f"column_{idx}"
            for idx, cell in enumerate(header, start=1)
        ]
        duplicate_headers = {
            header_name: count
            for header_name, count in Counter(normalized_headers).items()
            if count > 1
        }

        profile["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "first_rows": first_rows,
                "headers": normalized_headers,
                "duplicate_headers": duplicate_headers,
                "nonblank_by_col_sample": dict(nonblank_by_col),
                "type_counts_by_col_sample": {
                    str(col_idx): dict(counter)
                    for col_idx, counter in sorted(type_counts_by_col.items())
                },
            }
        )

    output_path.write_text(json.dumps(profile, indent=2), encoding="utf-8")
    print(json.dumps({"profile": str(output_path), "sheet_count": len(profile["sheets"])}, indent=2))


if __name__ == "__main__":
    main()

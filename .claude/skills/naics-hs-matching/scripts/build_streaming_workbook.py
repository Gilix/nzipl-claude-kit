from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEETS = [
    ("Overview", "overview.csv"),
    ("NAICS_Summary", "naics_summary.csv"),
    ("Country_Summary", "country_summary.csv"),
    ("Tech_Summary", "tech_summary.csv"),
    ("NAICS_HS_Map", "naics_hs_map.csv"),
    ("Firms_Clean", "firms_clean.csv"),
    ("Quality_Report", "quality_report.csv"),
    ("Data_Dictionary", "data_dictionary.csv"),
    ("Duplicate_Entities", "duplicate_entities.csv"),
]

TEXT_FIELDS = {
    "sp_entity_id",
    "naics6",
    "naics2",
    "naics3",
    "naics4",
    "hs_code",
    "final_naics6",
    "source_naics6",
    "source_naics6_nace",
}

HEADER_FILL = PatternFill("solid", fgColor="1A1B1E")
OVERVIEW_FILL = PatternFill("solid", fgColor="56A360")
HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Aptos", size=10, color="111827")
THIN_GRAY = Side(style="thin", color="D9DEE7")
HEADER_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def is_text_field(header: str) -> bool:
    h = header.lower()
    return (
        h in TEXT_FIELDS
        or h.endswith("_codes")
        or "naics_title" in h
        or "definition" in h
        or "source_workbook" in h
    )


def convert_cell(header: str, value: str) -> Any:
    if value == "":
        return None
    if is_text_field(header):
        return value
    if value == "True":
        return True
    if value == "False":
        return False
    if re.fullmatch(r"-?\d+(\.\d+)?", value):
        parsed = float(value)
        if parsed.is_integer():
            return int(parsed)
        return parsed
    return value


def width_for(header: str, sheet_name: str) -> int:
    h = header.lower()
    if sheet_name == "Overview" and h == "value":
        return 72
    if "business_description" in h:
        return 58
    if "description" in h or "definition" in h:
        return 52
    if "company_name" in h:
        return 36
    if "green_technologies" in h or "green_hs_codes" in h or "component_name" in h:
        return 34
    if "address" in h:
        return 28
    if "country" in h or "city" in h or "sector" in h or "status" in h:
        return 20
    if "technology" in h or "stage" in h or "role" in h:
        return 20
    if "metric" in h or "field" in h or "sheet" in h:
        return 22
    if "share" in h:
        return 12
    if "count" in h or "rows" in h or "sum" in h:
        return 14
    if "code" in h or "naics" in h or "hs" in h:
        return 14
    if "id" in h:
        return 14
    return 16


def number_format(header: str) -> str | None:
    h = header.lower()
    if "share" in h:
        return "0.0%"
    if "revenue" in h or "headcount" in h or "count" in h or "rows" in h or "sum" in h:
        return "#,##0"
    if "weight" in h:
        return "0.000"
    if "year" in h:
        return "0"
    return None


def header_cells(ws, headers: list[str], sheet_name: str) -> list[WriteOnlyCell]:
    cells = []
    fill = OVERVIEW_FILL if sheet_name == "Overview" else HEADER_FILL
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
        cells.append(cell)
    return cells


def body_cells(ws, headers: list[str], row: list[str], style_body: bool) -> list[Any]:
    values = [convert_cell(headers[idx], value) if idx < len(headers) else value for idx, value in enumerate(row)]
    if not style_body:
        return values
    cells = []
    for idx, value in enumerate(values):
        cell = WriteOnlyCell(ws, value=value)
        cell.font = BODY_FONT
        fmt = number_format(headers[idx])
        if fmt:
            cell.number_format = fmt
        cells.append(cell)
    return cells


def add_csv_sheet(wb: Workbook, csv_path: Path, sheet_name: str) -> dict[str, Any]:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        headers = next(reader)
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width_for(header, sheet_name)
        ws.append(header_cells(ws, headers, sheet_name))

        row_count = 1
        style_body = sheet_name != "Firms_Clean"
        for row in reader:
            ws.append(body_cells(ws, headers, row, style_body))
            row_count += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count}"
    return {"sheet": sheet_name, "rows": row_count, "columns": len(headers)}


def verify_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    error_count = 0
    for ws in wb.worksheets:
        headers: list[Any] = []
        row_count = 0
        column_count = 0
        sample_rows = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            column_count = max(column_count, len(row))
            if row_count == 1:
                headers = list(row)
                continue
            if sample_rows < 49:
                sample_rows += 1
                for value in row:
                    if isinstance(value, str) and value in error_tokens:
                        error_count += 1
        sheets[ws.title] = {
            "rows": row_count,
            "columns": column_count,
            "headers": headers,
            "sample_rows_checked": sample_rows,
        }
    return {"path": str(path), "sheets": sheets, "formula_error_tokens_in_samples": error_count}


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: build_streaming_workbook.py <csv_dir> <output.xlsx>")

    csv_dir = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=True)
    stats = []
    for sheet_name, csv_file in SHEETS:
        stats.append(add_csv_sheet(wb, csv_dir / csv_file, sheet_name))
        print(json.dumps(stats[-1]))

    wb.save(output_path)
    verification = verify_workbook(output_path)
    verification["build_stats"] = stats
    verification_path = output_path.with_suffix(".verification.json")
    verification_path.write_text(json.dumps(verification, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(output_path), "verification": str(verification_path)}, indent=2))


if __name__ == "__main__":
    main()

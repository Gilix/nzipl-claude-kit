from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("usage: inspect_output.py <workbook.xlsx>")

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        freeze_panes = getattr(ws, "freeze_panes", "read_only_unavailable")
        filter_ref = getattr(getattr(ws, "auto_filter", None), "ref", "read_only_unavailable")
        print(f"{ws.title}: freeze={freeze_panes}, filter={filter_ref}")
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            print("  ", row[: min(8, len(row))])


if __name__ == "__main__":
    main()
